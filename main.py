"""Check Insurance Risk - Backend FastAPI (Railway)"""
from datetime import datetime
from typing import Dict, List, Optional

import base64
import io
import json

import uvicorn
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from auth import create_access_token, verify_password
from database import execute_query, execute_transaction
from models import (
    DashboardStats,
    DecisionEnum,
    DecisionRequest,
    InfoSourceInfo,
    LoginRequest,
    LoginResponse,
    RiskCheckRequest,
    RiskCheckResponse,
    RiskAnalysisInfo,
    SourceTypeEnum,
    UserInfo,
)
from reporting import export_to_excel, generate_dashboard_charts, generate_pdf_report
from security import get_current_user, get_admin_user
from utils import calculate_risk_score, normalize_country, perform_matching
from seed_admin import seed_default_user

app = FastAPI(
    title="Check Insurance Risk API",
    description="Sistema de análise de risco para seguradoras",
    version="3.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def ensure_default_user():
    try:
        seed_default_user()
    except Exception as e:
        print(f"[startup] Erro ao garantir utilizador padrão: {e}")


@app.get("/")
async def root():
    return {
        "message": "Check Insurance Risk API",
        "status": "Online",
        "version": "3.0.0",
        "timestamp": datetime.utcnow().isoformat(),
    }


@app.post("/api/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    try:
        query = """
            SELECT id, username, email, password_hash, role, is_active,
                   last_login, created_at
            FROM users
            WHERE (username = %s OR email = %s) AND is_active = true
        """
        users = execute_query(query, (request.username, request.username))

        if not users:
            raise HTTPException(status_code=401, detail="Credenciais inválidas")

        user = users[0]

        if not verify_password(request.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Credenciais inválidas")

        execute_query(
            "UPDATE users SET last_login = NOW() WHERE id = %s",
            (user["id"],),
        )

        token = create_access_token(
            {
                "id": user["id"],
                "username": user["username"],
                "email": user["email"],
                "role": user["role"],
            }
        )

        return {
            "success": True,
            "token": token,
            "user": {
                "id": user["id"],
                "username": user["username"],
                "email": user["email"],
                "role": user["role"],
                "last_login": user["last_login"],
                "created_at": user["created_at"],
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro no login: {e}")
        raise HTTPException(status_code=500, detail="Erro interno")


@app.get("/api/auth/me", response_model=UserInfo)
async def get_me(current_user: UserInfo = Depends(get_current_user)):
    return current_user


@app.get("/api/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        total = execute_query("SELECT COUNT(*) AS count FROM risk_records")[0]["count"]

        pending = execute_query(
            """
            SELECT COUNT(*) AS count
            FROM risk_records
            WHERE decision = 'UNDER_REVIEW' OR decision IS NULL
            """
        )[0]["count"]

        high_risk = execute_query(
            """
            SELECT COUNT(*) AS count
            FROM risk_records
            WHERE risk_level IN ('HIGH', 'CRITICAL')
            """
        )[0]["count"]

        sources = execute_query(
            "SELECT COUNT(*) AS count FROM info_sources WHERE is_active = true"
        )[0]["count"]

        risk_dist_rows = execute_query(
            """
            SELECT risk_level, COUNT(*) AS count
            FROM risk_records
            WHERE risk_level IS NOT NULL
            GROUP BY risk_level
            """
        )
        risk_distribution: Dict[str, int] = {
            row["risk_level"]: row["count"] for row in risk_dist_rows
        }

        recent_rows = execute_query(
            """
            SELECT r.id, r.full_name, r.risk_level, r.risk_score,
                   r.analyzed_at, r.decision, u.username AS analyst_name
            FROM risk_records r
            LEFT JOIN users u ON r.analyzed_by = u.id
            ORDER BY r.analyzed_at DESC
            LIMIT 10
            """
        )

        recent: List[RiskAnalysisInfo] = []
        for row in recent_rows:
            recent.append(
                RiskAnalysisInfo(
                    id=row["id"],
                    full_name=row.get("full_name"),
                    risk_level=row.get("risk_level"),
                    risk_score=row.get("risk_score"),
                    analyzed_at=row.get("analyzed_at"),
                    decision=row.get("decision"),
                    analyst_name=row.get("analyst_name"),
                )
            )

        return {
            "totalAnalyses": total,
            "pendingReview": pending,
            "highRiskCases": high_risk,
            "activeSources": sources,
            "recentAnalyses": recent,
            "riskDistribution": risk_distribution,
        }
    except Exception as e:
        print(f"Erro dashboard: {e}")
        raise HTTPException(status_code=500, detail="Erro interno")


@app.post("/api/risk/check", response_model=RiskCheckResponse)
async def risk_check(
    request: RiskCheckRequest,
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        if not any(
            [
                request.full_name,
                request.nif,
                request.passport,
                request.resident_card,
            ]
        ):
            raise HTTPException(
                status_code=400,
                detail="Pelo menos um identificador é necessário",
            )

        search_country = None
        if request.nationality:
            search_country = normalize_country(request.nationality)

        matches = perform_matching(
            {
                "full_name": request.full_name,
                "nif": request.nif,
                "passport": request.passport,
                "resident_card": request.resident_card,
                "country": search_country,
            }
        )

        risk_data = calculate_risk_score(matches, bool(request.nif), search_country)

        query = """
            INSERT INTO risk_records (
                full_name, nif, passport, resident_card, notes,
                risk_score, risk_level, matches, risk_factors,
                analyzed_by, analyzed_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            RETURNING id, analyzed_at
        """

        result = execute_query(
            query,
            (
                request.full_name,
                request.nif,
                request.passport,
                request.resident_card,
                request.notes,
                risk_data["score"],
                risk_data["level"],
                json.dumps(matches),
                json.dumps(risk_data["factors"]),
                current_user.id,
            ),
        )

        record_id = result[0]["id"]
        analyzed_at = result[0]["analyzed_at"]

        return {
            "success": True,
            "id": record_id,
            "risk_score": risk_data["score"],
            "risk_level": risk_data["level"],
            "matches": matches,
            "risk_factors": risk_data["factors"],
            "analyzed_at": analyzed_at,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro análise: {e}")
        raise HTTPException(status_code=500, detail="Erro interno")


@app.get("/api/info-sources", response_model=List[InfoSourceInfo])
async def get_info_sources(
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        rows = execute_query(
            """
            SELECT s.*, u.username AS uploaded_by_name
            FROM info_sources s
            LEFT JOIN users u ON s.uploaded_by = u.id
            WHERE s.is_active = true
            ORDER BY s.uploaded_at DESC
            """
        )
        return rows
    except Exception as e:
        print(f"Erro fontes: {e}")
        raise HTTPException(status_code=500, detail="Erro interno")


@app.post("/api/info-sources/upload")
async def upload_info_source(
    name: str = Form(...),
    source_type: SourceTypeEnum = Form(...),
    file: UploadFile = File(...),
    current_user: UserInfo = Depends(get_admin_user),
):
    """
    Upload de fonte de informação em EXCEL (.xlsx) ou CSV.
    Sem suporte a PDF / URL por enquanto.
    """
    import csv
    from io import StringIO
    from openpyxl import load_workbook

    try:
        filename = file.filename or "fonte"
        content = await file.read()

        lower_name = filename.lower()
        rows: list[dict] = []

        if lower_name.endswith(".csv"):
            text = content.decode("utf-8", errors="ignore")
            sample = text.splitlines()[0] if text.splitlines() else ""
            delimiter = ";"
            if "," in sample and sample.count(",") >= sample.count(";"):
                delimiter = ","
            elif "\t" in sample:
                delimiter = "\t"
            reader = csv.DictReader(StringIO(text), delimiter=delimiter)
            rows = list(reader)

        elif lower_name.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active

            headers = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {
                    headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                    for i in range(len(headers))
                }
                rows.append(row_dict)
        else:
            raise HTTPException(
                status_code=400,
                detail="Tipo de ficheiro não suportado. Use .csv ou .xlsx",
            )

        def pick(d: dict, keys: list[str]) -> str | None:
            for k in keys:
                v = d.get(k)
                if v:
                    return str(v).strip()
            return None

        entities: list[dict] = []

        for row in rows:
            full_name = pick(
                row,
                ["Nome", "NOME", "Full Name", "FULL_NAME", "full_name", "name", "Name"],
            )
            nif = pick(row, ["NIF", "nif", "Tax ID", "tax_id", "NIF\ufeff"])
            position = pick(
                row,
                [
                    "Cargo",
                    "Função",
                    "Funcao",
                    "Position",
                    "role",
                    "ROLE",
                    "Função/Cargo",
                ],
            )
            country_raw = pick(
                row,
                [
                    "Nacionalidade",
                    "Nationality",
                    "País",
                    "Pais",
                    "Country",
                    "country",
                ],
            )

            country = normalize_country(country_raw) if country_raw else None

            if not full_name and not nif:
                continue

            entities.append(
                {
                    "full_name": full_name,
                    "nif": nif,
                    "position": position,
                    "country": country,
                    "additional_info": json.dumps(row, ensure_ascii=False),
                }
            )

        if not entities:
            raise HTTPException(
                status_code=400,
                detail="Nenhuma entidade válida encontrada no ficheiro",
            )

        src_rows = execute_query(
            """
            INSERT INTO info_sources (
                name, source_type, file_type, num_records,
                uploaded_at, uploaded_by, is_active
            )
            VALUES (%s, %s, %s, %s, NOW(), %s, true)
            RETURNING id
            """,
            (
                name,
                source_type.value,
                "EXCEL" if lower_name.endswith(".xlsx") else "CSV",
                len(entities),
                current_user.id,
            ),
        )
        source_id = src_rows[0]["id"]

        queries = []
        for ent in entities:
            queries.append(
                (
                    """
                    INSERT INTO normalized_entities (
                        full_name, nif, passport, resident_card,
                        position, country, additional_info, source_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        ent.get("full_name"),
                        ent.get("nif"),
                        None,
                        None,
                        ent.get("position"),
                        ent.get("country"),
                        ent.get("additional_info"),
                        source_id,
                    ),
                )
            )

        execute_transaction(queries)

        return {
            "success": True,
            "source_id": source_id,
            "inserted": len(entities),
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro upload fonte (Excel/CSV): {e}")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao processar fonte de informação",
        )


@app.get("/api/risk/{risk_id}/report/pdf")
async def download_risk_pdf(
    risk_id: int,
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        records = execute_query(
            "SELECT * FROM risk_records WHERE id = %s",
            (risk_id,),
        )
        if not records:
            raise HTTPException(status_code=404, detail="Registo não encontrado")

        risk_record = records[0]
        pdf_info = generate_pdf_report(risk_record)
        pdf_bytes = base64.b64decode(pdf_info["data"])
        filename = pdf_info.get("filename", f"risk_report_{risk_id}.pdf")

        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao gerar PDF: {e}")
        raise HTTPException(status_code=500, detail="Erro interno ao gerar PDF")


@app.get("/api/risk/export/excel")
async def export_risk_excel(
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        records = execute_query(
            """
            SELECT *
            FROM risk_records
            ORDER BY analyzed_at DESC
            """
        )

        excel_info = export_to_excel(records)
        excel_bytes = base64.b64decode(excel_info["data"])
        filename = excel_info.get(
            "filename",
            f"risk_analysis_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        raise HTTPException(status_code=500, detail="Erro interno ao exportar Excel")


@app.get("/api/dashboard/charts")
async def get_charts(
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        chart_info = generate_dashboard_charts()
        return chart_info
    except Exception as e:
        print(f"Erro ao gerar gráficos: {e}")
        raise HTTPException(status_code=500, detail="Erro interno ao gerar gráficos")


@app.put("/api/risk/{risk_id}/decision")
async def update_risk_decision(
    risk_id: int,
    request: DecisionRequest,
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        result = execute_query(
            """
            UPDATE risk_records
            SET decision = %s,
                analyst_notes = %s
            WHERE id = %s
            RETURNING id, full_name, risk_level, risk_score,
                      analyzed_at, decision, analyst_notes
            """,
            (request.decision.value, request.notes, risk_id),
        )

        if not result:
            raise HTTPException(status_code=404, detail="Registo não encontrado")

        return {"success": True, "record": result[0]}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao actualizar decisão: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao actualizar decisão de risco",
        )


@app.get("/api/risk/history")
async def get_risk_history(
    full_name: Optional[str] = None,
    nif: Optional[str] = None,
    passport: Optional[str] = None,
    resident_card: Optional[str] = None,
    current_user: UserInfo = Depends(get_current_user),
):
    try:
        if not any([full_name, nif, passport, resident_card]):
            raise HTTPException(
                status_code=400,
                detail="Pelo menos um identificador deve ser fornecido",
            )

        conditions = []
        params: List[str] = []

        if full_name:
            conditions.append("LOWER(full_name) LIKE LOWER(%s)")
            params.append(f"%{full_name}%")
        if nif:
            conditions.append("nif = %s")
            params.append(nif)
        if passport:
            conditions.append("passport = %s")
            params.append(passport)
        if resident_card:
            conditions.append("resident_card = %s")
            params.append(resident_card)

        where_clause = " OR ".join(conditions)

        query = f"""
            SELECT id, full_name, nif, passport, resident_card,
                   risk_level, risk_score, analyzed_at, decision
            FROM risk_records
            WHERE {where_clause}
            ORDER BY analyzed_at DESC
        """

        history = execute_query(query, tuple(params))

        return {"success": True, "count": len(history), "history": history}

    except HTTPException:
        raise
    except Exception as e:
        print(f"Erro ao obter histórico: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erro interno ao obter histórico do assegurado",
        )


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000)
