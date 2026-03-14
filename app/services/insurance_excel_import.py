from __future__ import annotations

import io
import uuid
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.models import InsurancePolicy, Payment, Claim, Cancellation, FraudFlag


def _s(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        return t if t else None
    t = str(v).strip()
    return t if t else None


def _upper(v: Any) -> Optional[str]:
    s = _s(v)
    return s.upper() if s else None


def _dt(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)

    s = _s(v)
    if not s:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def _int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _read_sheet_case_insensitive(wb, wanted_name: str) -> List[Dict[str, Any]]:
    real_name = None
    for s in wb.sheetnames:
        if str(s).strip().lower() == wanted_name.strip().lower():
            real_name = s
            break

    if not real_name:
        return []

    ws = wb[real_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if not any(x is not None and str(x).strip() != "" for x in r):
            continue
        item = {
            headers[i]: r[i]
            for i in range(min(len(headers), len(r)))
            if headers[i]
        }
        out.append(item)
    return out


def _read_active_records_fallback(wb) -> List[Dict[str, Any]]:
    ws = None
    for s in wb.sheetnames:
        if str(s).strip().lower() == "records":
            ws = wb[s]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []

    for r in rows[1:]:
        if not any(x is not None and str(x).strip() != "" for x in r):
            continue
        item = {
            headers[i]: r[i]
            for i in range(min(len(headers), len(r)))
            if headers[i]
        }
        out.append(item)
    return out


def _pick(row: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in row and row.get(k) is not None and str(row.get(k)).strip() != "":
            return row.get(k)
    return None


def _subject_fields(row: Dict[str, Any]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    full_name = _s(_pick(row, "subject_full_name", "full_name", "name"))
    bi = _s(_pick(row, "subject_bi", "bi", "id_number", "bi_number"))
    passport = _s(_pick(row, "subject_passport", "passport", "passport_number"))
    return full_name, bi, passport


def delete_previous_import(db: Session, *, entity_id: str, source_ref: str) -> None:
    for model in (InsurancePolicy, Payment, Claim, Cancellation, FraudFlag):
        db.query(model).filter(
            model.entity_id == entity_id,
            model.source_ref == source_ref,
        ).delete(synchronize_session=False)
    db.flush()


def import_insurance_workbook(
    db: Session,
    *,
    entity_id: str,
    source_name: str,
    source_ref: str,
    filename: str,
    content: bytes,
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(
        io.BytesIO(content),
        data_only=True,
        read_only=True,
    )

    policies_rows = _read_sheet_case_insensitive(wb, "policies")
    payments_rows = _read_sheet_case_insensitive(wb, "payments")
    claims_rows = _read_sheet_case_insensitive(wb, "claims")
    cancellations_rows = _read_sheet_case_insensitive(wb, "cancellations")
    fraud_rows = _read_sheet_case_insensitive(wb, "fraud_flags")

    if not any((policies_rows, payments_rows, claims_rows, cancellations_rows, fraud_rows)):
        policies_rows = _read_active_records_fallback(wb)

    inserted = {
        "policies": 0,
        "payments": 0,
        "claims": 0,
        "cancellations": 0,
        "fraud_flags": 0,
    }
    invalid: List[Dict[str, Any]] = []

    def add_policy(row: Dict[str, Any]) -> None:
        full_name, bi, passport = _subject_fields(row)
        product_type = _upper(_pick(row, "product_type", "produto", "product", "tipo_seguro"))
        if not product_type:
            raise ValueError("product_type obrigatório em policies")

        obj = InsurancePolicy(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            subject_full_name=full_name,
            subject_bi=bi,
            subject_passport=passport,
            product_type=product_type,
            policy_number=_s(_pick(row, "policy_number", "policy_no", "policy")),
            insurer_name=_s(_pick(row, "insurer_name", "seguradora", "insurer")),
            status=_upper(_pick(row, "status", "policy_status", "estado")),
            start_date=_dt(_pick(row, "start_date", "policy_start_date", "inicio")),
            end_date=_dt(_pick(row, "end_date", "policy_end_date", "fim")),
            currency=_upper(_pick(row, "currency", "moeda")),
            premium_amount=_int(_pick(row, "premium_amount", "premium", "premio")),
            sum_insured=_int(_pick(row, "sum_insured", "capital_segurado")),
            source_name=source_name,
            source_ref=source_ref,
            raw_payload=row,
        )
        db.add(obj)
        inserted["policies"] += 1

    def add_payment(row: Dict[str, Any]) -> None:
        full_name, bi, passport = _subject_fields(row)
        product_type = _upper(_pick(row, "product_type", "product", "tipo_seguro")) or "N/A"

        obj = Payment(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            subject_full_name=full_name,
            subject_bi=bi,
            subject_passport=passport,
            product_type=product_type,
            policy_number=_s(_pick(row, "policy_number", "policy_no", "policy")),
            amount=_int(_pick(row, "amount", "valor")),
            currency=_upper(_pick(row, "currency", "moeda")),
            paid_at=_dt(_pick(row, "paid_at", "paid_date", "data_pagamento")),
            due_at=_dt(_pick(row, "due_at", "due_date", "data_vencimento")),
            status=_upper(_pick(row, "status", "payment_status", "estado_pagamento")),
            source_name=source_name,
            source_ref=source_ref,
            raw_payload=row,
        )
        db.add(obj)
        inserted["payments"] += 1

    def add_claim(row: Dict[str, Any]) -> None:
        full_name, bi, passport = _subject_fields(row)
        product_type = _upper(_pick(row, "product_type", "product", "tipo_seguro")) or "N/A"

        obj = Claim(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            subject_full_name=full_name,
            subject_bi=bi,
            subject_passport=passport,
            product_type=product_type,
            policy_number=_s(_pick(row, "policy_number", "policy_no", "policy")),
            claim_number=_s(_pick(row, "claim_number", "sinistro_no", "claim_no")),
            loss_date=_dt(_pick(row, "loss_date", "claim_date", "data_sinistro")),
            reported_at=_dt(_pick(row, "reported_at", "data_report")),
            status=_upper(_pick(row, "status", "claim_status", "estado")),
            amount_claimed=_int(_pick(row, "amount_claimed", "valor_reclamado")),
            amount_paid=_int(_pick(row, "amount_paid", "claim_amount", "valor_pago")),
            currency=_upper(_pick(row, "currency", "moeda")),
            source_name=source_name,
            source_ref=source_ref,
            raw_payload=row,
        )
        db.add(obj)
        inserted["claims"] += 1

    def add_cancellation(row: Dict[str, Any]) -> None:
        full_name, bi, passport = _subject_fields(row)
        product_type = _upper(_pick(row, "product_type", "product", "tipo_seguro")) or "N/A"

        obj = Cancellation(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            subject_full_name=full_name,
            subject_bi=bi,
            subject_passport=passport,
            product_type=product_type,
            policy_number=_s(_pick(row, "policy_number", "policy_no", "policy")),
            cancelled_at=_dt(_pick(row, "cancelled_at", "date", "data_cancelamento")),
            reason=_s(_pick(row, "reason", "motivo")),
            source_name=source_name,
            source_ref=source_ref,
            raw_payload=row,
        )
        db.add(obj)
        inserted["cancellations"] += 1

    def add_fraud(row: Dict[str, Any]) -> None:
        full_name, bi, passport = _subject_fields(row)
        product_type = _upper(_pick(row, "product_type", "product", "tipo_seguro")) or "N/A"

        flag_type = _upper(_pick(row, "flag_type", "flag", "tipo_flag"))
        if not flag_type:
            raise ValueError("flag_type obrigatório em fraud_flags")

        obj = FraudFlag(
            id=str(uuid.uuid4()),
            entity_id=entity_id,
            subject_full_name=full_name,
            subject_bi=bi,
            subject_passport=passport,
            product_type=product_type,
            policy_number=_s(_pick(row, "policy_number", "policy_no", "policy")),
            flag_type=flag_type,
            severity=_upper(_pick(row, "severity", "severidade")),
            description=_s(_pick(row, "description", "note", "notes", "descricao")),
            source_name=source_name,
            source_ref=source_ref,
            raw_payload=row,
        )
        db.add(obj)
        inserted["fraud_flags"] += 1

    delete_previous_import(db, entity_id=entity_id, source_ref=source_ref)

    def ingest(rows: List[Dict[str, Any]], fn, key: str) -> None:
        for i, row in enumerate(rows, start=2):
            try:
                fn(row)
            except Exception as e:
                invalid.append({
                    "sheet": key,
                    "row_number": i,
                    "error": str(e),
                    "raw": row,
                })

    ingest(policies_rows, add_policy, "policies")
    ingest(payments_rows, add_payment, "payments")
    ingest(claims_rows, add_claim, "claims")
    ingest(cancellations_rows, add_cancellation, "cancellations")
    ingest(fraud_rows, add_fraud, "fraud_flags")

    return {
        "filename": filename,
        "inserted": inserted,
        "invalid": len(invalid),
        "invalid_rows": invalid[:30],
    }
