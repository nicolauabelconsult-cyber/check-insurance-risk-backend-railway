"""
info_sources.py
Gestão das Fontes de Informação (InfoSource) e importação de bases Excel
sem usar pandas, apenas openpyxl.
"""

from typing import List, Optional
from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    status,
    UploadFile,
    File,
    Form,
)
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from database import get_db
from models import InfoSource, NormalizedEntity, User
from schemas import InfoSourceRead
from auth import get_current_active_user, get_current_admin


router = APIRouter(prefix="/info-sources", tags=["Info Sources"])


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def _require_excel(filename: str) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="O ficheiro deve ser um Excel .xlsx",
        )


def _get_col_idx(headers: List[str], name: str) -> Optional[int]:
    name_norm = name.strip().lower()
    for idx, h in enumerate(headers):
        if not h:
            continue
        if str(h).strip().lower() == name_norm:
            return idx
    return None


def _str_or_none(value) -> Optional[str]:
    if value is None:
        return None
    v = str(value).strip()
    return v or None


# ------------------------------------------------------------
# 1) Listar fontes
# ------------------------------------------------------------

@router.get("/", response_model=List[InfoSourceRead])
def list_info_sources(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    sources = db.query(InfoSource).order_by(InfoSource.created_at.desc()).all()
    return sources


# ------------------------------------------------------------
# 2) Upload Excel
# ------------------------------------------------------------

@router.post("/upload-excel")
async def upload_info_source_excel(
    file: UploadFile = File(...),
    name: str = Form(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):

    if not file.filename:
        raise HTTPException(400, "Ficheiro inválido.")

    _require_excel(file.filename)

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Erro ao ler Excel: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(400, "O ficheiro está vazio.")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    idx_name = _get_col_idx(headers, "name")
    idx_nif = _get_col_idx(headers, "nif")
    idx_passport = _get_col_idx(headers, "passport")
    idx_res_card = _get_col_idx(headers, "resident_card")
    idx_country = _get_col_idx(headers, "country")

    if idx_name is None:
        raise HTTPException(400, "A coluna 'name' é obrigatória.")

    # Criar a fonte
    source = InfoSource(name=name, description=description)
    db.add(source)
    db.commit()
    db.refresh(source)

    # Criar entidades normalizadas
    created = 0
    for row in rows[1:]:
        if not row:
            continue

        full_name = _str_or_none(row[idx_name])
        if not full_name:
            continue

        entity = NormalizedEntity(
            name=full_name,
            normalized_name=full_name.upper(),
            nif=_str_or_none(row[idx_nif]) if idx_nif is not None else None,
            passport=_str_or_none(row[idx_passport]) if idx_passport is not None else None,
            resident_card=_str_or_none(row[idx_res_card]) if idx_res_card is not None else None,
            country=(_str_or_none(row[idx_country]).upper()
                     if idx_country is not None and _str_or_none(row[idx_country])
                     else None),
            info_source_id=source.id,
        )

        db.add(entity)
        created += 1

    db.commit()

    return {
        "message": "Ficheiro importado com sucesso.",
        "source_id": source.id,
        "registos_importados": created,
    }


# ------------------------------------------------------------
# 3) Amostra de entidades
# ------------------------------------------------------------

@router.get("/{source_id}/sample")
def sample_entities(
    source_id: int,
    limit: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    source = db.query(InfoSource).filter(InfoSource.id == source_id).first()
    if not source:
        raise HTTPException(404, "Fonte não encontrada.")

    entities = (
        db.query(NormalizedEntity)
        .filter(NormalizedEntity.info_source_id == source_id)
        .order_by(NormalizedEntity.created_at.desc())
        .limit(limit)
        .all()
    )

    return {
        "source": {
            "id": source.id,
            "name": source.name,
            "description": source.description,
            "created_at": source.created_at,
        },
        "sample": [
            {
                "id": e.id,
                "name": e.name,
                "normalized_name": e.normalized_name,
                "nif": e.nif,
                "passport": e.passport,
                "resident_card": e.resident_card,
                "country": e.country,
            }
            for e in entities
        ],
    }
